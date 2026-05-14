"""
SQL Generator UI
================
A desktop interface to:
  1. Upload a mapping Excel file
  2. Read the Control Sheet and display all sheets with checkboxes
  3. Select which tables/sheets to process
  4. Generate & validate SQL and export to Excel report
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# CORE LOGIC (same as sql_v2.py)
# ─────────────────────────────────────────────

col_map_keys = {
    "source_table":   ["Table Name", "Source Table", "Raw Table"],
    "source_column":  ["Alias", "Source Column", "Raw Column", "Technical Name"],
    "transformation": ["Transformations", "Transformation", "SQL Logic"],
    "target_table":   ["Table name", "Target Table"],
    "target_column":  ["Field name", "Target Column", "Curated Column"]
}


# ❌ DO NOT CHANGE
def build_automation_script(table_name, s1_sql, t1_sql, output_path):
    return f"""
from datetime import date, datetime
# Initialize settings
settings_instance = Settings()
settings_instance.setsparkobject(spark)
settings_instance.projectinfo("RISE EDL", "raw")
settings_instance.setoutputpath("{output_path}")
settings_instance.projectname = "RISE EDL"
print(len(settings_instance.projectname))
settings_instance.projectdesc = "Auto Generated"
settings_instance.setwriteformat("delta")
settings_instance.setmaxerror(-1)
settings_instance.setloglevel("all")
settings_instance.replacelist = {{}}

# Initialize Writer
init_writer()

# Create TestRunner
tr = TestRunner(settings_instance)

for i in range(1):
    # Set table property
    tr.settableproperty(
        tablename="{table_name}",
        coltest=True,
        coldatatypetest=True,
        datatest=True,
        savetable=False,
        debug=True
        #excolsincoltest=["load_dt", "load_ts"],
        #excolsindatatest=["load_dt", "load_ts"]
    )
    # Initialize test
    tr.inittest()
    # Load source --- QA SCRIPT
    s1 = tr.loadsql(\"\"\"\n{s1_sql}\n    \"\"\", "source")
    # Load target
    t1 = tr.loadsql(\"\"\"\n{t1_sql}\n    \"\"\", "target")
    
    # Build tables with correct PK
    s2 = tr.buildtable("src1", "SELECT * FROM source", pk="generated_primary_key")
    t2 = tr.buildtable("trg1", "SELECT * FROM target", pk="generated_primary_key")
    # Column test
    ctr = tr.comparecolumns("source", "target")
    
    # Key test
    ktr = tr.comparekeycolumn(s2, t2)

    if ktr and ktr.get("MATCH_FOUND", 0) > 0:
        dtr = tr.comparedata(s2, t2)
        print("Data Test:", dtr)
    else:
        print("No matches found to do data testing.")

    # Finish table
    tsr = tr.finishtable()
    print("Table Summary:", tsr)

print("test completed....")

sr = tr.getsummary()
print(sr)

repname = tr.finishtest()
print(repname)

tr.printsampletestreport()
"""



def build_sql_expression(src_col, tgt_col, transform):
    src_col   = str(src_col).strip()
    tgt_col_str = "" if pd.isna(tgt_col) else str(tgt_col).strip()
    transform   = "" if pd.isna(transform) else str(transform).strip()

    if not transform or transform.upper() == "DIRECT":
        if str(src_col).strip().lower() in ["", "null", "none", "nan", "na"]:
            return f"NULL AS {tgt_col_str}"
        return f"{src_col} AS {tgt_col_str}"

    parts = [p.strip().lower() for p in transform.split('+')]

    if "direct" in parts or "audit column" in parts:
        base_expr = src_col
    elif "explode_outer" in parts:
        base_expr = f"explode_outer({src_col})"
    elif "single select from value list" in parts:
        base_expr = f"array_join({src_col}, ', ')"
    else:
        base_expr = transform

    if "convert to string" in parts:
        base_expr = f"CAST({base_expr} AS STRING)"

    if tgt_col_str and re.search(rf"\bAS\s+{re.escape(tgt_col_str)}\b", str(base_expr), re.IGNORECASE):
        return base_expr

    return f"{base_expr} AS {tgt_col_str}" if tgt_col_str else base_expr




# ─────────────────────────────────────────────
# MAIN ENGINE
# ─────────────────────────────────────────────

def validate_sql(meta):
    sql            = meta["sql"]
    tgt_columns    = meta["tgt_columns"]
    src_columns    = meta["src_columns"]
    transforms     = meta["transforms"]
    select_clauses = meta["select_clauses"]

    findings, scores, column_issues = [], {}, {}

    # Check 1: SELECT (10 pts)
    scores["SELECT_clause"] = 10 if re.search(r"\bSELECT\b", sql, re.IGNORECASE) else 0
    if not scores["SELECT_clause"]:
        findings.append("❌ [SELECT] Missing SELECT keyword.")

    # Check 2: FROM (10 pts)
    fm = re.search(r"\bFROM\s+(\S+)", sql, re.IGNORECASE)
    scores["FROM_clause"] = 10 if fm and fm.group(1) not in ("", "None", "null", "nan") else 0
    if not scores["FROM_clause"]:
        findings.append("❌ [FROM] Missing or null source table in FROM clause.")

    # Check 3: AS aliases (15 pts)
    alias_issues = []
    for i, clause in enumerate(select_clauses):
        if "AS" not in clause.upper() and "NO_MAPPINGS" not in clause:
            col_name = tgt_columns[i] if i < len(tgt_columns) else f"Column_{i}"
            alias_issues.append(col_name)
            column_issues.setdefault(col_name, []).append("Missing AS alias")
    scores["AS_aliases"] = max(0, 15 - min(15, len(alias_issues) * 3))
    if alias_issues:
        findings.append(f"⚠️ [ALIAS] {len(alias_issues)} column(s) missing AS alias: {', '.join(alias_issues[:5])}")

    # Check 4: Column count (10 pts)
    scores["column_count"] = 10 if tgt_columns else 0
    if not scores["column_count"]:
        findings.append("❌ [COLUMNS] No target columns found.")

    # Check 5: Balanced parentheses (15 pts)
    o, c = sql.count("("), sql.count(")")
    scores["parentheses"] = 15 if o == c else 0
    if o != c:
        findings.append(f"❌ [SYNTAX] Unbalanced parentheses: {o} '(' vs {c} ')'.")

    # Check 6: Transformations (20 pts)
    ts, t_issues = 20, []
    for i, (tr, cl) in enumerate(zip(transforms, select_clauses)):
        tl, cu = tr.lower(), cl.upper()
        cn = tgt_columns[i] if i < len(tgt_columns) else f"Column_{i}"
        if ("array_join" in tl or "single select" in tl) and "ARRAY_JOIN" not in cu:
            ts -= 2; t_issues.append(f"{cn} (Expected ARRAY_JOIN)")
            column_issues.setdefault(cn, []).append("Missing ARRAY_JOIN transformation")
        if "explode" in tl and "EXPLODE" not in cu:
            ts -= 2; t_issues.append(f"{cn} (Expected EXPLODE)")
            column_issues.setdefault(cn, []).append("Missing EXPLODE transformation")
        if "convert to string" in tl and "CAST" not in cu:
            ts -= 2; t_issues.append(f"{cn} (Expected CAST AS STRING)")
            column_issues.setdefault(cn, []).append("Missing CAST AS STRING transformation")
    scores["transformations"] = max(0, ts)
    if t_issues:
        findings.append(f"⚠️ [TRANSFORM] {len(t_issues)} issue(s): {', '.join(t_issues[:5])}")

    # Check 7: No NaN/None literals (10 pts)
    nan_hits = re.findall(r"\bNaN\b|\bNone\b", sql)
    nan_cols = []
    for i, cl in enumerate(select_clauses):
        if re.search(r"\bNaN\b|\bNone\b", cl):
            cn = tgt_columns[i] if i < len(tgt_columns) else f"Column_{i}"
            nan_cols.append(cn)
            column_issues.setdefault(cn, []).append("Contains NaN/None literal")
    scores["no_null_literals"] = max(0, 10 - len(nan_hits) * 2) if nan_hits else 10
    if nan_hits:
        findings.append(f"⚠️ [NULL] {len(nan_hits)} NaN/None literal(s) in: {', '.join(nan_cols[:5])}")

    # Check 8: Empty source columns (10 pts)
    empty_src = [tgt for src, tgt in zip(src_columns, tgt_columns)
                 if str(src).strip().lower() in ("", "nan", "none")]
    for tgt in empty_src:
        column_issues.setdefault(tgt, []).append("Empty source column")
    scores["src_columns_populated"] = max(0, 10 - min(10, len(empty_src) * 2)) if empty_src else 10
    if empty_src:
        findings.append(f"⚠️ [SOURCE] {len(empty_src)} column(s) have empty source: {', '.join(empty_src[:5])}")

    total = sum(scores.values())
    grade = "EXCELLENT" if total >= 90 else "GOOD" if total >= 75 else "FAIR" if total >= 55 else "POOR"

    if not findings:
        findings.append("✅ No issues found — SQL looks clean!")

    return {"total_score": total, "grade": grade, "breakdown": scores,
            "findings": findings, "column_issues": column_issues}


def run_generation(file_name, sheets_to_process, log_fn, done_fn):
    """Core generation logic — runs in a background thread."""
    try:
        all_sql_results    = []
        all_sheet_metadata = []

        for mapping_sheet_name in sheets_to_process:
            log_fn(f"\n📄 Processing sheet: {mapping_sheet_name}")

            header_row = None
            for i in range(15):
                temp_df = pd.read_excel(file_name, sheet_name=mapping_sheet_name, header=i)
                cols = [str(c).strip() for c in temp_df.columns]
                if sum(any(c in cols for c in v) for v in col_map_keys.values()) >= 3:
                    header_row = i
                    break

            if header_row is None:
                log_fn(f"  ⚠️ Could not detect header row — skipping.")
                continue

            pdf_mapping = pd.read_excel(file_name, sheet_name=mapping_sheet_name, header=header_row)

            COL_MAP = {}
            for key, names in col_map_keys.items():
                found = next((c for c in names if c in pdf_mapping.columns), None)
                if not found:
                    log_fn(f"  ⚠️ Column for '{key}' not found — skipping sheet.")
                    break
                COL_MAP[key] = found
            else:
                # Only runs if inner loop didn't break
                select_clauses, tgt_columns, src_columns, transforms_used = [], [], [], []
                source_table = None

                for _, row in pdf_mapping.iterrows():
                    src_col   = row[COL_MAP["source_column"]]
                    tgt_col   = row[COL_MAP["target_column"]]
                    transform = row[COL_MAP["transformation"]]
                    source_table = row[COL_MAP["source_table"]]

                    if pd.isna(tgt_col) or str(tgt_col).strip() == "":
                        continue

                    expr = build_sql_expression(src_col, tgt_col, transform)
                    select_clauses.append(f"    {expr}")
                    tgt_columns.append(str(tgt_col).strip() if not pd.isna(tgt_col) else "")
                    src_columns.append(str(src_col).strip() if not pd.isna(src_col) else "")
                    transforms_used.append(str(transform).strip() if not pd.isna(transform) else "DIRECT")

                if not select_clauses:
                    continue

                s1_sql = (
                    "SELECT\n    " +
                    ",\n    ".join(select_clauses) +
                    "\n    ,row_number() OVER (ORDER BY load_dt DESC) AS generated_primary_key\n"
                    f"FROM {source_table}"
                )

                log_fn("\n" + "=" * 120)
                log_fn(f"AUTOMATION SCRIPT FOR TABLE: {source_table}")
                log_fn("=" * 120)

                log_fn("\n--- Load source --- QA SCRIPT SQL (Projection Query)")
                log_fn("-" * 120)
                log_fn(s1_sql)
                log_fn("-" * 120)

                atom_script = None

                generate_atom = True  # Assuming this flag is needed; set True or False as required

                if generate_atom:
                    output_path = (
                        "abfss://gf-risk-landing@cactgftriskmsci01adls.dfs.core.windows.net/"
                        f"Raise_EDL/Habibur/MAY/Semantic/{source_table}_"
                        f"{datetime.now().strftime('%d%m%Y')}_Test_Result"
                    )

                    t1_sql = (
                        "select\n    " +
                        ",\n    ".join(tgt_columns) +
                        f"\nfrom {source_table}"
                    )

                    atom_script = build_automation_script(source_table, s1_sql, t1_sql, output_path)
                    log_fn(atom_script)

                all_sheet_metadata.append({
                    "sheet": mapping_sheet_name,
                    "source_table": source_table,
                    "tgt_columns": tgt_columns,
                    "src_columns": src_columns,
                    "transforms": transforms_used,
                    "select_clauses": select_clauses,
                    "sql": s1_sql,
                    "atom_script": atom_script
                })

        # ── Validation ──────────────────────────────────────────────────────
        log_fn("\n" + "="*60)
        log_fn("🤖 VALIDATION REPORT")
        log_fn("="*60)

        validation_results = []
        all_column_issues  = {}

        for meta in all_sheet_metadata:
            result = validate_sql(meta)
            validation_results.append({
                "Sheet": meta["sheet"], "Source_Table": meta["source_table"],
                "Columns_Mapped": len(meta["tgt_columns"]),
                "Total_Score": result["total_score"], "Grade": result["grade"],
                "Findings": " | ".join(result["findings"]), "SQL": meta["sql"]
            })
            for col_name, issues in result["column_issues"].items():
                all_column_issues[f"{meta['sheet']}|{col_name}"] = issues

            log_fn(f"\n📋 {meta['sheet']}  |  Score: {result['total_score']}/100  |  {result['grade']}")
            for f in result["findings"]:
                log_fn(f"   {f}")

        if not validation_results:
            log_fn("\n⚠️ No sheets were processed successfully.")
            done_fn(None)
            return

        # ── Excel Export ─────────────────────────────────────────────────────
        overall_avg = sum(r["Total_Score"] for r in validation_results) / len(validation_results)

        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"SQL_Generation_Report_{timestamp}.xlsx"
        input_dir   = os.path.dirname(os.path.abspath(file_name))
        output_path = os.path.join(input_dir, output_file)

        df_summary = pd.DataFrame(validation_results)

        detail_rows = []
        for meta in all_sheet_metadata:
            for src, tgt, tr, cl in zip(meta["src_columns"], meta["tgt_columns"],
                                         meta["transforms"], meta["select_clauses"]):
                key = f"{meta['sheet']}|{tgt}"
                issues_text = " | ".join(all_column_issues.get(key, []))
                detail_rows.append({
                    "Sheet_Name": meta["sheet"], "Source_Table": meta["source_table"],
                    "Source_Column": src, "Target_Column": tgt,
                    "Transformation": tr, "Generated_Clause": cl.strip(),
                    "Issues": issues_text if issues_text else "No Issues"
                })
        df_detail = pd.DataFrame(detail_rows)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            df_detail.to_excel(writer, sheet_name='Column_Detail', index=False)
            df_summary[['Sheet', 'Source_Table', 'Columns_Mapped', 'SQL']].to_excel(
                writer, sheet_name='Generated_SQL', index=False)

        # ── Styling ───────────────────────────────────────────────────────────
        wb = load_workbook(output_path)

        def grade_fill(score):
            if score >= 90: return "C6EFCE", "276221"
            if score >= 75: return "FFEB9C", "9C6500"
            if score >= 55: return "FFCC99", "974706"
            return "FFC7CE", "9C0006"

        def style_sheet(ws, header_color):
            hf  = PatternFill("solid", fgColor=header_color)
            hfn = Font(bold=True, color="FFFFFF", size=11)
            bs  = Side(style="thin", color="BFBFBF")
            bdr = Border(left=bs, right=bs, top=bs, bottom=bs)
            for cell in ws[1]:
                cell.fill = hf; cell.font = hfn
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = bdr
            gi = si = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "Grade":      gi = idx
                if cell.value == "Total_Score": si = idx
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.border = bdr
                if gi and si:
                    sv = row[si - 1].value
                    if isinstance(sv, (int, float)):
                        bg, fg = grade_fill(sv)
                        row[gi - 1].fill = PatternFill("solid", fgColor=bg)
                        row[gi - 1].font = Font(color=fg, bold=True)
            for col_cells in ws.columns:
                ml = max((len(str(c.value)) if c.value else 0 for c in col_cells), default=10)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(ml + 4, 80)
            ws.freeze_panes = "A2"

        # Cover page
        wsc = wb.create_sheet("Cover", 0)
        wsc["B2"] = "SQL Generation & Validation Report"
        wsc["B2"].font = Font(bold=True, size=20, color="1F4E79")

        wsc["B4"] = "📋 Run Details";  wsc["B4"].font = Font(bold=True, size=13, color="1F4E79")
        details = [
            ("Generated On:",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Input File:",          os.path.basename(file_name)),
            ("Sheets Processed:",    len(all_sheet_metadata)),
            ("Total Columns Mapped:",sum(len(m["tgt_columns"]) for m in all_sheet_metadata)),
            ("Overall Avg Score:",   f"{overall_avg:.1f}/100"),
        ]
        for r, (label, value) in enumerate(details, start=6):
            wsc[f"B{r}"] = label; wsc[f"B{r}"].font = Font(bold=True)
            wsc[f"C{r}"] = value
        bg, fg = grade_fill(overall_avg)
        wsc["C10"].fill = PatternFill("solid", fgColor=bg)
        wsc["C10"].font = Font(color=fg, bold=True)

        status_map = {90: "🟢 Production Ready", 75: "🟡 Minor Issues – Review Required",
                      55: "🟠 Fair – Fixes Needed",  0: "🔴 Critical Issues – Must Rework"}
        status = next(v for k, v in sorted(status_map.items(), reverse=True) if overall_avg >= k)
        wsc["B11"] = "Status:"; wsc["B11"].font = Font(bold=True)
        wsc["C11"] = status;    wsc["C11"].font = Font(color=fg, bold=True)

        wsc["B13"] = "📊 Sheet Summary"; wsc["B13"].font = Font(bold=True, size=13, color="1F4E79")
        hdr_row = 15
        for col, header in zip(["B","C","D","E","F"],
                                ["Sheet Name","Source Table","Columns","Score","Grade"]):
            c = wsc[f"{col}{hdr_row}"]
            c.value = header; c.font = Font(bold=True, color="FFFFFF")
            c.fill  = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))

        for rn, res in enumerate(validation_results, start=hdr_row + 1):
            for col, val in zip(["B","C","D","E","F"],
                                 [res["Sheet"], res["Source_Table"], res["Columns_Mapped"],
                                  res["Total_Score"], res["Grade"]]):
                cell = wsc[f"{col}{rn}"]
                cell.value = val
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                     top=Side(style="thin"), bottom=Side(style="thin"))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            gbg, gfg = grade_fill(res["Total_Score"])
            wsc[f"F{rn}"].fill = PatternFill("solid", fgColor=gbg)
            wsc[f"F{rn}"].font = Font(color=gfg, bold=True)

        end_row = hdr_row + len(validation_results) + 3
        wsc[f"B{end_row}"] = "📄 Report Contents"
        wsc[f"B{end_row}"].font = Font(bold=True, size=13, color="1F4E79")
        contents = ["1. Cover – This page",
                    "2. Summary – Scores and findings per sheet",
                    "3. Column_Detail – Row-level mapping with issues highlighted",
                    "4. Generated_SQL – Full SQL queries"]
        for i, line in enumerate(contents, start=end_row + 2):
            wsc[f"B{i}"] = line
        wsc.column_dimensions["B"].width = 28
        wsc.column_dimensions["C"].width = 42
        wsc.column_dimensions["D"].width = 12
        wsc.column_dimensions["E"].width = 12
        wsc.column_dimensions["F"].width = 22

        style_sheet(wb["Summary"],      "1F4E79")
        style_sheet(wb["Generated_SQL"],"7B2C2C")

        # Column_Detail with row highlighting
        wsd = wb["Column_Detail"]
        hdr_fill = PatternFill("solid", fgColor="375623")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        bs  = Side(style="thin", color="BFBFBF")
        bdr = Border(left=bs, right=bs, top=bs, bottom=bs)
        for cell in wsd[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
        issues_idx = next((idx for idx, c in enumerate(wsd[1], 1) if c.value == "Issues"), None)
        for row in wsd.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.border = bdr
            if issues_idx:
                ic = row[issues_idx - 1]
                if ic.value and ic.value != "No Issues":
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFE6E6")
                    ic.fill = PatternFill("solid", fgColor="FFC7CE")
                    ic.font = Font(color="9C0006", bold=True)
        for col_cells in wsd.columns:
            ml = max((len(str(c.value)) if c.value else 0 for c in col_cells), default=10)
            wsd.column_dimensions[get_column_letter(col_cells[0].column)].width = min(ml + 4, 80)
        wsd.freeze_panes = "A2"

        # ── ATOM_Script worksheet ─────────────────────────────────────────────
        ws_atom = wb.create_sheet("ATOM_Script")

        # Header
        ws_atom["A1"] = "ATOM Automation Scripts"
        ws_atom["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        ws_atom["A1"].fill = PatternFill("solid", fgColor="1F3864")
        ws_atom["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_atom.merge_cells("A1:C1")
        ws_atom.row_dimensions[1].height = 30

        ws_atom["A2"] = "Generated On:"
        ws_atom["A2"].font = Font(bold=True)
        ws_atom["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        current_row = 4
        atom_hdr_fill = PatternFill("solid", fgColor="1F3864")
        atom_sep_fill = PatternFill("solid", fgColor="D6E4F0")
        mono_font     = Font(name="Courier New", size=9)

        for meta in all_sheet_metadata:
            script_text = meta.get("atom_script", "")
            if not script_text:
                continue

            # Section header for each table
            ws_atom.cell(row=current_row, column=1, value=f"▶  Table: {meta['source_table']}  |  Sheet: {meta['sheet']}")
            hdr_cell = ws_atom.cell(row=current_row, column=1)
            hdr_cell.font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_cell.fill  = atom_hdr_fill
            hdr_cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws_atom.merge_cells(start_row=current_row, start_column=1,
                                end_row=current_row, end_column=3)
            ws_atom.row_dimensions[current_row].height = 22
            current_row += 1

            # ── Entire script in ONE cell ──────────────────────────────────
            script_cell = ws_atom.cell(row=current_row, column=1, value=script_text.strip())
            script_cell.font      = mono_font
            script_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws_atom.merge_cells(start_row=current_row, start_column=1,
                                end_row=current_row, end_column=3)
            # Set row height based on number of lines so full script is visible
            line_count = script_text.count("\n") + 1
            ws_atom.row_dimensions[current_row].height = max(15 * line_count, 100)
            current_row += 1

            # Separator row
            for col in range(1, 4):
                sep = ws_atom.cell(row=current_row, column=col, value="")
                sep.fill = atom_sep_fill
            ws_atom.merge_cells(start_row=current_row, start_column=1,
                                end_row=current_row, end_column=3)
            ws_atom.row_dimensions[current_row].height = 8
            current_row += 2

        ws_atom.column_dimensions["A"].width = 120
        ws_atom.freeze_panes = "A3"

        # Update Cover page contents list to include ATOM_Script
        atom_contents_row = end_row + 2 + len(contents)
        wsc[f"B{atom_contents_row}"] = "5. ATOM_Script – Full automation scripts for ATOM test runner"

        wb.save(output_path)

        log_fn(f"\n{'='*60}")
        log_fn(f"✅ Report saved: {output_path}")
        log_fn(f"🏁 Overall Average Score: {overall_avg:.1f}/100  |  {status}")
        log_fn("="*60)

        done_fn(output_path)

    except Exception as e:
        import traceback
        log_fn(f"\n❌ ERROR: {e}\n{traceback.format_exc()}")
        done_fn(None)


# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────

class SqlGeneratorApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("SQL Generator & Validator")
        self.geometry("900x700")
        self.minsize(750, 580)
        self.configure(bg="#F4F6F9")
        self.resizable(True, True)

        self._file_path    = tk.StringVar()
        self._sheet_vars   = {}   # sheet_name -> BooleanVar
        self._running      = False

        self._build_ui()

    # ── UI construction ──────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header bar ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg="#1F4E79", height=56)
        hdr.pack(fill="x")
        tk.Label(hdr, text="  🗂  SQL Generator & Validator",
                 font=("Segoe UI", 16, "bold"), fg="white", bg="#1F4E79",
                 anchor="w").pack(side="left", pady=10, padx=10)

        # ── Main body ─────────────────────────────────────────────────────────
        body = tk.Frame(self, bg="#F4F6F9")
        body.pack(fill="both", expand=True, padx=18, pady=14)

        # Step 1: File selection
        self._section(body, "Step 1 — Upload Mapping File", row=0)
        file_frame = tk.Frame(body, bg="#F4F6F9")
        file_frame.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        body.columnconfigure(0, weight=1)

        self._file_entry = tk.Entry(file_frame, textvariable=self._file_path,
                                    font=("Segoe UI", 10), width=60,
                                    relief="solid", bd=1)
        self._file_entry.pack(side="left", fill="x", expand=True, ipady=4)
        self._browse_btn = self._btn(file_frame, "Browse…", self._browse_file,
                                     bg="#1F4E79", fg="white")
        self._browse_btn.pack(side="left", padx=(8, 0))

        # Step 2: Sheet selection
        self._section(body, "Step 2 — Select Tables to Process", row=2)

        sheet_outer = tk.Frame(body, bg="#FFFFFF", relief="solid", bd=1)
        sheet_outer.grid(row=3, column=0, sticky="ew", pady=(0, 12))

        # Toolbar: Select All / Deselect All / Load
        tb = tk.Frame(sheet_outer, bg="#EBF0F7")
        tb.pack(fill="x", padx=0, pady=0)
        self._btn(tb, "⟳ Load Sheets", self._load_sheets,
                  bg="#375623", fg="white").pack(side="left", padx=6, pady=5)
        self._btn(tb, "☑ Select All",   self._select_all,
                  bg="#555", fg="white").pack(side="left", padx=4, pady=5)
        self._btn(tb, "☐ Deselect All", self._deselect_all,
                  bg="#555", fg="white").pack(side="left", padx=4, pady=5)

        # Scrollable checkbox area
        canvas_frame = tk.Frame(sheet_outer, bg="#FFFFFF")
        canvas_frame.pack(fill="both", expand=True)
        self._canvas = tk.Canvas(canvas_frame, bg="#FFFFFF", height=160,
                                 highlightthickness=0)
        vsb = ttk.Scrollbar(canvas_frame, orient="vertical",
                            command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._check_frame = tk.Frame(self._canvas, bg="#FFFFFF")
        self._canvas_window = self._canvas.create_window(
            (0, 0), window=self._check_frame, anchor="nw")
        self._check_frame.bind("<Configure>", self._on_frame_configure)
        self._canvas.bind("<Configure>", self._on_canvas_configure)
        self._placeholder = tk.Label(self._check_frame,
            text="← Browse for a file then click  ⟳ Load Sheets",
            font=("Segoe UI", 9, "italic"), fg="#888", bg="#FFFFFF")
        self._placeholder.grid(row=0, column=0, padx=12, pady=12, sticky="w")

        # Step 3: Run
        self._section(body, "Step 3 — Generate & Export", row=4)
        run_frame = tk.Frame(body, bg="#F4F6F9")
        run_frame.grid(row=5, column=0, sticky="ew", pady=(0, 12))
        self._run_btn = self._btn(run_frame, "▶  Generate SQL + Validate + Export Excel",
                                  self._run, bg="#7B2C2C", fg="white",
                                  font=("Segoe UI", 11, "bold"))
        self._run_btn.pack(side="left")
        self._open_btn = self._btn(run_frame, "📂 Open Report", self._open_report,
                                   bg="#375623", fg="white")
        self._open_btn.pack(side="left", padx=(12, 0))
        self._open_btn.config(state="disabled")
        self._output_path = None

        # Log area
        self._section(body, "Output Log", row=6)
        self._log = scrolledtext.ScrolledText(
            body, font=("Consolas", 9), bg="#1A1A2E", fg="#E0E0FF",
            insertbackground="white", relief="solid", bd=1, height=14,
            wrap="word", state="disabled")
        self._log.grid(row=7, column=0, sticky="nsew", pady=(0, 8))
        body.rowconfigure(7, weight=1)

        # Progress bar
        self._progress = ttk.Progressbar(body, mode="indeterminate")
        self._progress.grid(row=8, column=0, sticky="ew")

        # ── Tag colours for log ───────────────────────────────────────────────
        self._log.tag_config("ok",   foreground="#7FFF7F")
        self._log.tag_config("warn", foreground="#FFD700")
        self._log.tag_config("err",  foreground="#FF6B6B")
        self._log.tag_config("info", foreground="#87CEEB")

    # ── Helpers ────────────────────────────────────────────────────────────

    @staticmethod
    def _btn(parent, text, cmd, bg="#1F4E79", fg="white",
             font=("Segoe UI", 9, "bold")):
        b = tk.Button(parent, text=text, command=cmd,
                      bg=bg, fg=fg, font=font,
                      relief="flat", cursor="hand2",
                      activebackground=bg, activeforeground=fg,
                      padx=10, pady=5)
        return b

    @staticmethod
    def _section(parent, text, row):
        lbl = tk.Label(parent, text=text,
                       font=("Segoe UI", 10, "bold"), fg="#1F4E79", bg="#F4F6F9")
        lbl.grid(row=row, column=0, sticky="w", pady=(6, 2))

    def _on_frame_configure(self, _):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self._canvas.itemconfig(self._canvas_window, width=event.width)

    # ── Actions ──────────────────────────────────────────────────────────────

    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Mapping Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._file_path.set(path)
            self._load_sheets()

    def _load_sheets(self):
        path = self._file_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("No File", "Please browse and select a valid Excel file first.")
            return

        try:
            # Try reading Control_Sheet
            try:
                df = pd.read_excel(path, sheet_name="Control_Sheet",
                                   usecols=["Sheet_Name", "Run_flag"])
                df["Run_flag"] = df["Run_flag"].astype(str).str.strip().str.upper()
                sheets = df[["Sheet_Name", "Run_flag"]].dropna(subset=["Sheet_Name"]).values.tolist()
            except Exception:
                # Fallback: list all sheet names
                xl = pd.ExcelFile(path)
                sheets = [(s, "Y") for s in xl.sheet_names if s != "Control_Sheet"]

            # Clear existing checkboxes
            for w in self._check_frame.winfo_children():
                w.destroy()
            self._sheet_vars.clear()

            if not sheets:
                tk.Label(self._check_frame, text="No sheets found.",
                         font=("Segoe UI", 9, "italic"), fg="#888",
                         bg="#FFFFFF").grid(row=0, column=0, padx=12, pady=8)
                return

            # Column headers
            tk.Label(self._check_frame, text="Include", font=("Segoe UI", 9, "bold"),
                     bg="#EBF0F7", width=8, anchor="center").grid(
                row=0, column=0, padx=(8, 2), pady=4, sticky="w")
            tk.Label(self._check_frame, text="Sheet / Table Name",
                     font=("Segoe UI", 9, "bold"), bg="#EBF0F7", anchor="w").grid(
                row=0, column=1, padx=4, pady=4, sticky="w")
            tk.Label(self._check_frame, text="Run Flag",
                     font=("Segoe UI", 9, "bold"), bg="#EBF0F7", width=10).grid(
                row=0, column=2, padx=4, pady=4, sticky="w")

            for i, (sheet_name, run_flag) in enumerate(sheets, start=1):
                var = tk.BooleanVar(value=(str(run_flag).strip().upper() == "Y"))
                self._sheet_vars[str(sheet_name)] = var

                bg_row = "#FFFFFF" if i % 2 == 0 else "#F8F9FC"
                cb = tk.Checkbutton(self._check_frame, variable=var,
                                    bg=bg_row, activebackground=bg_row)
                cb.grid(row=i, column=0, padx=(12, 2), pady=2, sticky="w")

                tk.Label(self._check_frame, text=str(sheet_name),
                         font=("Segoe UI", 9), bg=bg_row, anchor="w",
                         width=38).grid(row=i, column=1, padx=4, pady=2, sticky="w")

                flag_color = "#276221" if str(run_flag).upper() == "Y" else "#9C0006"
                tk.Label(self._check_frame, text=str(run_flag),
                         font=("Segoe UI", 9, "bold"), fg=flag_color,
                         bg=bg_row, width=8).grid(row=i, column=2, padx=4, pady=2, sticky="w")

            self._log_write(f"✅ Loaded {len(sheets)} sheet(s) from control sheet.", "ok")

        except Exception as e:
            messagebox.showerror("Error Loading File", str(e))

    def _select_all(self):
        for var in self._sheet_vars.values():
            var.set(True)

    def _deselect_all(self):
        for var in self._sheet_vars.values():
            var.set(False)

    def _run(self):
        if self._running:
            return

        path = self._file_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("No File", "Please select a valid Excel file.")
            return

        selected = [name for name, var in self._sheet_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("No Sheets Selected",
                                   "Please select at least one sheet/table to process.")
            return

        self._running = True
        self._run_btn.config(state="disabled", text="⏳ Running…")
        self._open_btn.config(state="disabled")
        self._output_path = None
        self._progress.start(10)

        self._log_clear()
        self._log_write(f"🚀 Starting — {len(selected)} sheet(s) selected", "info")
        self._log_write(f"   File: {path}", "info")
        self._log_write(f"   Sheets: {', '.join(selected)}\n", "info")

        threading.Thread(
            target=run_generation,
            args=(path, selected, self._log_write_safe, self._on_done),
            daemon=True
        ).start()

    def _on_done(self, output_path):
        self._progress.stop()
        self._running = False
        self._run_btn.config(state="normal", text="▶  Generate SQL + Validate + Export Excel")
        if output_path:
            self._output_path = output_path
            self._open_btn.config(state="normal")
            messagebox.showinfo("Done", f"Report saved!\n\n{output_path}")
        else:
            messagebox.showerror("Failed", "Generation failed. Check the Output Log for details.")

    def _open_report(self):
        if self._output_path and os.path.exists(self._output_path):
            os.startfile(self._output_path)

    # ── Logging ──────────────────────────────────────────────────────────────

    def _log_write(self, msg, tag=None):
        self._log.config(state="normal")
        if tag is None:
            if msg.startswith("✅") or msg.startswith("🚀") or msg.startswith("🏁"):
                tag = "ok"
            elif "❌" in msg or "ERROR" in msg:
                tag = "err"
            elif "⚠️" in msg:
                tag = "warn"
            else:
                tag = "info"
        self._log.insert("end", msg + "\n", tag)
        self._log.see("end")
        self._log.config(state="disabled")

    def _log_write_safe(self, msg):
        """Thread-safe log write via after()."""
        self.after(0, self._log_write, msg)

    def _log_clear(self):
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    app = SqlGeneratorApp()
    app.mainloop()